"""Excel-Export aller Deals mit allen Fakten (Konzept: eine private Übersicht,
die die Excel-Datei vollständig ersetzt - der Export ist der Weg zurück
in ein Tabellenformat, z. B. für die Steuererklärung oder ein Backup)."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import derived
from .models import Deal


def _autosize(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 60)


def _header(ws: Worksheet, titel: list[str]) -> None:
    ws.append(titel)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def build_workbook(deals: list[Deal]) -> io.BytesIO:
    wb = Workbook()

    deals_ws = wb.active
    deals_ws.title = "Deals"
    _header(
        deals_ws,
        [
            "ID",
            "Bank",
            "Kontoart",
            "Inhaber",
            "Status",
            "Kontonummer",
            "Kündbar ab",
            "Gekündigt",
            "Gekündigt im Monat",
            "Kündigung bestätigt",
            "Kündigung Anweisungen",
            "Kündigung Link",
            "Freibetrag",
            "Prämien auf Sparkonto",
            "Zugangsdaten gesichert",
            "Kommentar",
            "Prämien gesamt",
            "Prämien erhalten",
            "Prämien offen",
            "Erstellt am",
            "Geändert am",
        ],
    )
    for d in deals:
        kz = derived.kennzahlen(d.praemien)
        deals_ws.append(
            [
                d.id,
                d.bank.name,
                d.kontoart,
                d.inhaber.name,
                derived.STATUS_LABELS[derived.status(d)],
                d.kontonummer,
                d.kuendbar_ab,
                "Ja" if d.gekuendigt else "Nein",
                d.gekuendigt_im_monat,
                "Ja" if d.kuendigung_bestaetigt else "Nein",
                d.kuendigung_hinweis,
                d.kuendigung_hinweis_url,
                float(d.freibetrag) if d.freibetrag is not None else None,
                {True: "Ja", False: "Nein", None: None}[d.praemien_auf_sparkonto],
                "Ja" if d.zugangsdaten_gespeichert else "Nein",
                d.kommentar,
                float(kz.gesamt),
                float(kz.erhalten),
                float(kz.offen),
                d.erstellt_am.strftime("%d.%m.%Y %H:%M") if d.erstellt_am else None,
                d.geaendert_am.strftime("%d.%m.%Y %H:%M") if d.geaendert_am else None,
            ]
        )
    _autosize(deals_ws)

    praemien_ws = wb.create_sheet("Prämien")
    _header(praemien_ws, ["Deal-ID", "Bank", "Kontoart", "Inhaber", "Quelle", "Betrag", "Erhalten", "Auszahlung erwartet"])
    for d in deals:
        for p in d.praemien:
            praemien_ws.append(
                [d.id, d.bank.name, d.kontoart, d.inhaber.name, p.quelle, float(p.betrag), "Ja" if p.erhalten else "Nein", p.auszahlung_erwartet]
            )
    _autosize(praemien_ws)

    bedingungen_ws = wb.create_sheet("Bedingungen")
    _header(bedingungen_ws, ["Deal-ID", "Bank", "Kontoart", "Inhaber", "Beschreibung", "Erfüllt", "Fällig bis"])
    for d in deals:
        for b in d.bedingungen:
            bedingungen_ws.append(
                [d.id, d.bank.name, d.kontoart, d.inhaber.name, b.beschreibung, "Ja" if b.erfuellt else "Nein", b.faellig_bis]
            )
    _autosize(bedingungen_ws)

    aufgaben_ws = wb.create_sheet("Aufgaben")
    _header(aufgaben_ws, ["Deal-ID", "Bank", "Kontoart", "Inhaber", "Beschreibung", "Erledigt", "Fällig bis"])
    for d in deals:
        for a in d.aufgaben:
            aufgaben_ws.append(
                [d.id, d.bank.name, d.kontoart, d.inhaber.name, a.beschreibung, "Ja" if a.erledigt else "Nein", a.faellig_bis]
            )
    _autosize(aufgaben_ws)

    urls_ws = wb.create_sheet("Links")
    _header(urls_ws, ["Deal-ID", "Bank", "Kontoart", "Inhaber", "Bezeichnung", "URL"])
    for d in deals:
        for u in d.urls:
            urls_ws.append([d.id, d.bank.name, d.kontoart, d.inhaber.name, u.bezeichnung, u.url])
    _autosize(urls_ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
